import customtkinter as ctk
from tkinter import messagebox
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import os

# ==================== CONFIGURAÇÕES ====================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

CAMINHO_PLANILHA = r"C:\Users\Dell\Documents\CURSODEPYTONFABIO\CRM_MARCENARIA\estudos_26\cadastro de clientes\cadastro de cliente.xlsx"
NOME_ABA = "Clientes"


# =======================================================
# Função auxiliar para abrir planilha
# =======================================================
def abrir_planilha():
    try:
        wb = load_workbook(CAMINHO_PLANILHA)
        ws = wb[NOME_ABA]
        return wb, ws
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível abrir a planilha:\n{e}")
        return None, None


# =======================================================
# CADASTRAR CLIENTE
# =======================================================
def salvar_cliente():
    wb, ws = abrir_planilha()
    if ws is None:
        return

    nome = entry_nome.get().strip().upper()
    telefone = entry_telefone.get().strip()
    email = entry_email.get().strip()
    endereco = entry_endereco.get().strip()
    servico = entry_servico.get().strip()
    valor = entry_valor.get().strip()
    pagamento = entry_pagamento.get().strip()

    if nome == "":
        messagebox.showwarning("Aviso", "O nome não pode ficar vazio.")
        return

    ws.append([nome, telefone, email, endereco, servico, valor, pagamento])
    wb.save(CAMINHO_PLANILHA)

    messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")

    # limpar campos
    entry_nome.delete(0, "end")
    entry_telefone.delete(0, "end")
    entry_email.delete(0, "end")
    entry_endereco.delete(0, "end")
    entry_servico.delete(0, "end")
    entry_valor.delete(0, "end")
    entry_pagamento.delete(0, "end")


# =======================================================
# EDITAR CLIENTE
# =======================================================
def editar_cliente():
    wb, ws = abrir_planilha()
    if ws is None:
        return

    nome_busca = entry_busca_editar.get().strip().upper()

    for row in ws.iter_rows(min_row=2):
        if row[0].value and row[0].value.upper() == nome_busca:

            row[0].value = entry_edit_nome.get().strip() or row[0].value
            row[1].value = entry_edit_telefone.get().strip() or row[1].value
            row[2].value = entry_edit_email.get().strip() or row[2].value
            row[3].value = entry_edit_endereco.get().strip() or row[3].value
            row[4].value = entry_edit_servico.get().strip() or row[4].value
            row[5].value = entry_edit_valor.get().strip() or row[5].value
            row[6].value = entry_edit_pagamento.get().strip() or row[6].value

            wb.save(CAMINHO_PLANILHA)
            messagebox.showinfo("Sucesso", "Cliente editado com sucesso!")
            return

    messagebox.showerror("Erro", "Cliente não encontrado.")


# =======================================================
# EXCLUIR CLIENTE
# =======================================================
def excluir_cliente():
    wb, ws = abrir_planilha()
    if ws is None:
        return

    nome_busca = entry_busca_excluir.get().strip().upper()

    for row in ws.iter_rows(min_row=2):
        if row[0].value and row[0].value.upper() == nome_busca:
            resposta = messagebox.askyesno("Confirmar", "Tem certeza que deseja excluir?")
            if resposta:
                ws.delete_rows(row[0].row)
                wb.save(CAMINHO_PLANILHA)
                messagebox.showinfo("Sucesso", "Cliente excluído com sucesso!")
            return

    messagebox.showerror("Erro", "Cliente não encontrado.")


# =======================================================
# LISTAR CLIENTES + PDF
# =======================================================
def listar_clientes():
    wb, ws = abrir_planilha()
    if ws is None:
        return

    textbox_listar.delete("0.0", "end")

    for row in ws.iter_rows(min_row=1, values_only=True):
        if any(row):
            linha = " | ".join([str(c) for c in row])
            textbox_listar.insert("end", linha + "\n")


def exportar_pdf():
    wb, ws = abrir_planilha()
    if ws is None:
        return

    pdf_path = "clientes_exportados.pdf"
    cnv = canvas.Canvas(pdf_path, pagesize=A4)

    y = 800
    cnv.setFont("Helvetica", 12)
    cnv.drawString(50, y, "LISTA DE CLIENTES")
    y -= 30

    for row in ws.iter_rows(min_row=1, values_only=True):
        if any(row):
            texto = " | ".join([str(c) for c in row])
            cnv.drawString(50, y, texto)
            y -= 20
            if y < 50:
                cnv.showPage()
                y = 800

    cnv.save()
    messagebox.showinfo("PDF Gerado", f"Arquivo criado:\n{os.path.abspath(pdf_path)}")


# ==================== INTERFACE ====================
app = ctk.CTk()
app.title("CRM - Cadastro de Clientes")
app.geometry("850x600")

tabview = ctk.CTkTabview(app)
tabview.pack(expand=True, fill="both")

tab_cadastro = tabview.add("Cadastrar Cliente")
tab_editar = tabview.add("Editar Cliente")
tab_excluir = tabview.add("Excluir Cliente")
tab_listar = tabview.add("Listar Clientes")


# =======================================================
# ABA CADASTRAR
# =======================================================
ctk.CTkLabel(tab_cadastro, text="Cadastrar Novo Cliente", font=("Arial", 18)).pack(pady=10)

entry_nome = ctk.CTkEntry(tab_cadastro, placeholder_text="Nome")
entry_telefone = ctk.CTkEntry(tab_cadastro, placeholder_text="Telefone")
entry_email = ctk.CTkEntry(tab_cadastro, placeholder_text="Email")
entry_endereco = ctk.CTkEntry(tab_cadastro, placeholder_text="Endereço")
entry_servico = ctk.CTkEntry(tab_cadastro, placeholder_text="Serviço")
entry_valor = ctk.CTkEntry(tab_cadastro, placeholder_text="Valor")
entry_pagamento = ctk.CTkEntry(tab_cadastro, placeholder_text="Forma de Pagamento")

widgets = [entry_nome, entry_telefone, entry_email, entry_endereco,
           entry_servico, entry_valor, entry_pagamento]

for w in widgets:
    w.pack(pady=5, fill="x", padx=20)

ctk.CTkButton(tab_cadastro, text="Salvar Cliente", command=salvar_cliente).pack(pady=20)


# =======================================================
# ABA EDITAR
# =======================================================
ctk.CTkLabel(tab_editar, text="Editar Cliente", font=("Arial", 18)).pack(pady=10)

entry_busca_editar = ctk.CTkEntry(tab_editar, placeholder_text="Nome do Cliente para Editar")
entry_busca_editar.pack(pady=5, fill="x", padx=20)

entry_edit_nome = ctk.CTkEntry(tab_editar, placeholder_text="Novo Nome (opcional)")
entry_edit_telefone = ctk.CTkEntry(tab_editar, placeholder_text="Novo Telefone")
entry_edit_email = ctk.CTkEntry(tab_editar, placeholder_text="Novo Email")
entry_edit_endereco = ctk.CTkEntry(tab_editar, placeholder_text="Novo Endereço")
entry_edit_servico = ctk.CTkEntry(tab_editar, placeholder_text="Novo Serviço")
entry_edit_valor = ctk.CTkEntry(tab_editar, placeholder_text="Novo Valor")
entry_edit_pagamento = ctk.CTkEntry(tab_editar, placeholder_text="Novo Pagamento")

for w in [
    entry_edit_nome, entry_edit_telefone, entry_edit_email,
    entry_edit_endereco, entry_edit_servico, entry_edit_valor,
    entry_edit_pagamento
]:
    w.pack(pady=5, fill="x", padx=20)

ctk.CTkButton(tab_editar, text="Salvar Alterações", command=editar_cliente).pack(pady=20)


# =======================================================
# ABA EXCLUIR
# =======================================================
ctk.CTkLabel(tab_excluir, text="Excluir Cliente", font=("Arial", 18)).pack(pady=10)

entry_busca_excluir = ctk.CTkEntry(tab_excluir, placeholder_text="Nome do Cliente para Excluir")
entry_busca_excluir.pack(pady=10, fill="x", padx=20)

ctk.CTkButton(tab_excluir, text="Excluir Cliente", fg_color="red", command=excluir_cliente).pack(pady=20)


# =======================================================
# ABA LISTAR + PDF
# =======================================================
ctk.CTkLabel(tab_listar, text="Lista de Clientes", font=("Arial", 18)).pack(pady=10)

textbox_listar = ctk.CTkTextbox(tab_listar, width=700, height=350)
textbox_listar.pack(pady=10)

ctk.CTkButton(tab_listar, text="Atualizar Lista", command=listar_clientes).pack(pady=5)
ctk.CTkButton(tab_listar, text="Exportar para PDF", command=exportar_pdf).pack(pady=10)


app.mainloop()
