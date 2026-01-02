
from openpyxl import load_workbook

caminho = r"C:\Users\Dell\Documents\CURSODEPYTONFABIO\CRM_MARCENARIA\estudos_26\cadastro de clientes\cadastro de cliente.xlsx"

planilhas_vendas = load_workbook(caminho)
df = planilhas_vendas.active

for linha in df.iter_rows(values_only=True):
    # remove os Nones do fim
    if all(c is None for c in linha):
        continue
    dados = [cel for cel in linha if cel is not None]
    #print(dados)

def CADASTRO_DE_CLIENTE ():
    pass
    from openpyxl import load_workbook

from openpyxl import load_workbook
#=======================================FUNÇÃO DE CADASTRO DE CLIENTE======================================
def CADASTRO_DE_CLIENTE():
    # Caminho correto da planilha Excel (.xlsx)
    caminho = r"C:\Users\Dell\Documents\CURSODEPYTONFABIO\CRM_MARCENARIA\estudos_26\cadastro de clientes\cadastro de cliente.xlsx"

    # Abrir planilha
    wb = load_workbook(caminho)
    ws = wb.active

    # Coletar dados
    nome = input("DIGITE O NOME DO CLIENTE: ").upper(). strip()
    telefone = input("DIGITE O TELEFONE DO CLIENTE: ")
    email = input("DIGITE O EMAIL DO CLIENTE: ")
    endereco = input("DIGITE O ENDEREÇO DO CLIENTE: ")
    servico = input("DIGITE O SERVIÇO DO CLIENTE: ")
    valor = input("DIGITE O VALOR DO SERVIÇO DO CLIENTE: ")
    pagamento = input("DIGITE O PAGAMENTO DO SERVIÇO DO CLIENTE: ")


    # Confirmar
    salvar = input("DESEJA SALVAR O CLIENTE? (S/N): ").upper()

    if salvar == "S":
        ws.append([nome, telefone, email, endereco, servico, valor, pagamento])
        wb.save(caminho)
        print("OK! CLIENTE SALVO COM SUCESSO!")
    else:
        print("CADASTRO CANCELADO.")

from openpyxl import load_workbook
#=======================================FUNÇÃO DE EDITAÇÃO DE CLIENTE======================================
def EDITAR_CLIENTE():
    caminho = r"C:\Users\Dell\Documents\CURSODEPYTONFABIO\CRM_MARCENARIA\estudos_26\cadastro de clientes\cadastro de cliente.xlsx"
    
    wb = load_workbook(caminho)
    ws = wb.active

    nome_busca = input("DIGITE O NOME DO CLIENTE: ").strip()

    cliente_encontrado = False

    # Procurar cliente linha por linha
    for row in ws.iter_rows(min_row=2):  # pula cabeçalho
        if row[0].value and row[0].value.strip().lower() == nome_busca.lower():

            cliente_encontrado = True

            print("\nCLIENTE ENCONTRADO!")
            print("DEIXE EM BRANCO PARA NÃO ALTERAR.\n")

            novo_nome = input(f"Nome ({row[0].value}): ") or row[0].value
            novo_telefone = input(f"Telefone ({row[1].value}): ") or row[1].value
            novo_email = input(f"Email ({row[2].value}): ") or row[2].value
            novo_endereco = input(f"Endereço ({row[3].value}): ") or row[3].value
            novo_servico = input(f"Serviço ({row[4].value}): ") or row[4].value
            novo_valor = input(f"Valor ({row[5].value}): ") or row[5].value
            novo_pagamento = input(f"Pagamento ({row[6].value}): ") or row[6].value

            # Atualizar a linha
            row[0].value = novo_nome
            row[1].value = novo_telefone
            row[2].value = novo_email
            row[3].value = novo_endereco
            row[4].value = novo_servico
            row[5].value = novo_valor
            row[6].value = novo_pagamento

            wb.save(caminho)

            print("\n✔ CLIENTE EDITADO COM SUCESSO!\n")
            break

    if not cliente_encontrado:
        print("❌ CLIENTE NÃO ENCONTRADO.")


from openpyxl import load_workbook
#=======================================FUNÇÃO DE EXCLUSÃO DE CLIENTE======================================
def EXCLUIR_CLIENTE():
    caminho = r"C:\Users\Dell\Documents\CURSODEPYTONFABIO\CRM_MARCENARIA\estudos_26\cadastro de clientes\cadastro de cliente.xlsx"
    
    wb = load_workbook(caminho)
    ws = wb.active

    nome_busca = input("DIGITE O NOME DO CLIENTE PARA EXCLUIR: ").strip()

    cliente_encontrado = False

    # Percorrer a planilha para encontrar o cliente
    for row in ws.iter_rows(min_row=2):  # pula cabeçalho
        if row[0].value and row[0].value.strip().lower() == nome_busca.lower():

            cliente_encontrado = True
            numero_linha = row[0].row  # número REAL da linha no Excel

            print("\nCLIENTE ENCONTRADO:")
            print(f"Nome: {row[0].value}")
            print(f"Telefone: {row[1].value}")
            print(f"Email: {row[2].value}")
            print(f"Endereço: {row[3].value}")
            print(f"Serviço: {row[4].value}")
            print(f"Valor: {row[5].value}")
            print(f"Pagamento: {row[6].value}")

            confirmar = input("\nTEM CERTEZA QUE DESEJA EXCLUIR? (S/N): ").upper()

            if confirmar == "S":
                ws.delete_rows(numero_linha)  # exclui a linha inteira
                wb.save(caminho)
                print("\n✔ CLIENTE EXCLUÍDO COM SUCESSO!\n")
            else:
                print("\n❌ EXCLUSÃO CANCELADA.\n")

            break

    if not cliente_encontrado:
        print("❌ CLIENTE NÃO ENCONTRADO.")
#====================================== FUNÇÃO DE PRODUTOS PRESTADOS ================================
def PRODUTOS_PRESTADOS ():
    pass

# --- MENU ---
while True:
    print("DIGITE A OPÇÃO DESEJADA:")
    print("[1] - CADASTRO DE CLIENTE")
    print("[2] - EDITAR CLIENTE")
    print("[3] - EXCLUIR CLIENTE")
    print("[4] - SAIR")

    op = input(str("OPÇÂO: "))
    if op == "1": CADASTRO_DE_CLIENTE()
    elif op == "2":EDITAR_CLIENTE ()
    elif op == "3":EXCLUIR_CLIENTE()
    elif op == "4": break

